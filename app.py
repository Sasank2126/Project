from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
app.secret_key = "secret123"

EXCEL_FILE = "registration_data.xlsx"

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/register", methods=["POST"])
def register():
    name = request.form.get("name")
    email = request.form.get("email")
    number = request.form.get("number")
    gender = request.form.get("gender")
    blood_group = request.form.get("bloodGroup")
    age = request.form.get("age")
    location = request.form.get("location")

    # # Example other operations
    # print("Name:", name)
    # print("Email:", email)
    # print("Phone:", number)
    # print("Gender:", gender)
    # print("Blood Group:", blood_group)
    # print("Age:", age)
    # print("Location:", location)

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Name", "Email", "Phone", "Gender", "Blood Group", "Age", "Location"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    ws.append([name, email, number, gender, blood_group, age, location])
    wb.save(EXCEL_FILE)

    flash("Registration successful and data saved.")
    return redirect(url_for("home"))

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)